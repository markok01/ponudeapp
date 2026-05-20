from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import ExportMode
from app.utils.data_types import infer_column_type, parse_date, parse_number
from app.utils.logging_config import get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def generate_excel(
    tables: list[dict[str, Any]],
    output_path: Path,
    *,
    export_mode: ExportMode,
    base_name: str,
) -> tuple[int, int]:
    """Generate formatted Excel workbook. Returns (row_count, sheet_count)."""
    wb = Workbook()
    wb.remove(wb.active)

    prepared = _prepare_tables(tables, export_mode)

    if not prepared:
        ws = wb.create_sheet("Empty")
        ws.append(["No data extracted"])
        wb.save(output_path)
        return 0, 1

    total_rows = 0
    for sheet_name, table in prepared:
        safe_name = _safe_sheet_name(sheet_name)[:31]
        ws = wb.create_sheet(safe_name)
        _write_table(ws, table)
        total_rows += len(table["rows"])

    wb.save(output_path)
    logger.info("Excel saved: %s (%s sheets, %s rows)", output_path, len(prepared), total_rows)
    return total_rows, len(prepared)


def _prepare_tables(tables: list[dict[str, Any]], export_mode: ExportMode) -> list[tuple[str, dict[str, Any]]]:
    if export_mode == ExportMode.MULTIPLE_SHEETS:
        return [(t["name"], t) for t in tables if t.get("rows")]

    if export_mode == ExportMode.SINGLE_SHEET:
        combined_rows: list[list[Any]] = []
        columns: list[str] = []
        for i, table in enumerate(tables):
            if not table.get("rows"):
                continue
            if i > 0:
                combined_rows.append([])
                combined_rows.append([f"--- {table['name']} (page {table['page']}) ---"])
            if not columns and table.get("columns"):
                columns = table["columns"]
            combined_rows.extend(table["rows"])
        if not combined_rows:
            return []
        return [("All_Data", {"name": "All_Data", "page": 1, "columns": columns or ["Data"], "rows": combined_rows})]

    # COMBINED: merge tables with matching column structures
    groups: dict[str, dict[str, Any]] = {}
    for table in tables:
        if not table.get("rows"):
            continue
        key = "|".join(table.get("columns", []))
        if key not in groups:
            groups[key] = deepcopy(table)
            groups[key]["name"] = f"Combined_{len(groups)}"
        else:
            groups[key]["rows"].extend(table["rows"])

    return [(t["name"], t) for t in groups.values()]


def _write_table(ws, table: dict[str, Any]) -> None:
    columns: list[str] = table.get("columns") or []
    rows: list[list[Any]] = table.get("rows") or []

    if columns:
        ws.append(columns)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.freeze_panes = "A2"

    col_types = [infer_column_type([str(r[i]) if i < len(r) else "" for r in rows]) for i in range(len(columns))]

    for row_idx, row in enumerate(rows, start=2 if columns else 1):
        padded = list(row) + [""] * max(0, len(columns) - len(row))
        ws.append(padded[: len(columns)] if columns else padded)
        for col_idx, value in enumerate(padded[: len(columns)] if columns else padded, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            text = str(value) if value is not None else ""
            col_type = col_types[col_idx - 1] if col_idx - 1 < len(col_types) else "text"
            if col_type == "currency" or col_type == "number":
                num = parse_number(text)
                if num is not None:
                    cell.value = num
                    cell.number_format = '#,##0.00 "RSD"' if "RSD" in text.upper() else "#,##0.00"
            elif col_type == "date":
                dt = parse_date(text)
                if dt:
                    cell.value = dt
                    cell.number_format = "DD.MM.YYYY"

    _auto_column_width(ws)


def _auto_column_width(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 10)


def _safe_sheet_name(name: str) -> str:
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        name = name.replace(ch, "_")
    return name or "Sheet"
