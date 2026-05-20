from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HORECA_BRAND_FILL = PatternFill("solid", fgColor="B09FC6")
HORECA_HEADER_FILL = PatternFill("solid", fgColor="CCC0DA")
WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FONT = Font(bold=True, size=10)
PRODUCT_FONT = Font(size=10, bold=False)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(top=THIN, left=THIN, bottom=THIN, right=THIN)


def _col_count(layout: str) -> int:
    return 7 if layout == "wide" else 6


def _write_brand_row(ws, row_num: int, label: str, col_count: int) -> None:
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.fill = HORECA_BRAND_FILL
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    ws.row_dimensions[row_num].height = 18


def _is_price_literal(val: str) -> bool:
    v = (val or "").strip()
    if not v:
        return False
    return bool(
        re.search(r"din\s*/\s*(kom|kg|seta|l|ml)", v, re.I)
        or re.search(r"rsd\s*/\s*(kom|kg|seta|l|ml)", v, re.I)
        or re.search(r"^\d[\d.,\s]*\s*(din|rsd)\s*/", v, re.I)
    )


def _is_pdv_literal(val: str) -> bool:
    v = (val or "").strip()
    return bool(re.match(r"^(10|20)\s*%$", v)) or v in ("0.1", "0.2")


def _write_literal_cell(cell, val: str, *, align_right: bool = False) -> None:
    cell.value = val if val else ""
    cell.number_format = "@"
    if align_right:
        cell.alignment = Alignment(horizontal="right", vertical="top")


def _write_row(ws, row_num: int, cells: list[str], kind: str, layout: str) -> None:
    col_count = _col_count(layout)
    price_col = 6 if layout == "wide" else 5
    pdv_col = 7 if layout == "wide" else 6

    if kind == "brand":
        label = cells[0] if cells else ""
        _write_brand_row(ws, row_num, label, col_count)
        return

    for i in range(col_count):
        val = cells[i] if i < len(cells) else ""
        col = i + 1
        cell = ws.cell(row=row_num, column=col)
        cell.border = BORDER

        if _is_price_literal(str(val)):
            _write_literal_cell(cell, str(val), align_right=True)
        elif col == pdv_col and _is_pdv_literal(str(val)):
            _write_literal_cell(cell, str(val).strip(), align_right=True)
        else:
            cell.value = val

        if kind == "header":
            cell.fill = HORECA_HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif kind == "product":
            cell.font = PRODUCT_FONT
            if not _is_price_literal(str(val)) and col != pdv_col:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_horeca_workbook(sheets: list[dict[str, Any]], output_path: Path) -> tuple[int, int]:
    wb = Workbook()
    wb.remove(wb.active)

    total_rows = 0
    widths = [10, 12, 42, 42, 14, 18, 8]

    for sheet_data in sheets:
        name = str(sheet_data.get("name", "Sheet"))[:31]
        ws = wb.create_sheet(name)
        rows = sheet_data.get("rows", [])
        layout = sheet_data.get("layout", "compact")

        row_num = 0
        for row_data in rows:
            if row_data.get("kind") == "raw" and not any(row_data.get("cells", [])):
                continue
            row_num += 1
            _write_row(
                ws,
                row_num,
                row_data.get("cells", []),
                row_data.get("kind", "raw"),
                row_data.get("layout", layout),
            )
            total_rows += 1

        for idx, w in enumerate(widths[: _col_count(layout)], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    if not wb.sheetnames:
        ws = wb.create_sheet("Cenovnik")
        total_rows = 1

    wb.save(output_path)
    return total_rows, len(wb.sheetnames)
